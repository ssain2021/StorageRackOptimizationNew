
# Normal Comment
# ! Very Important Errors or things which must be fixed or be in attention immediately
# TODO: Things remaining to do
# ? Questions
# * Some Messages or Notes
# ^ Important Notes, Messages or things which may need attention in future
# & Important Commented out Code
# @ Section Starter
# ~ Run Time, PsuedoCode, Important Logic, Other Things

# TODO: Make some Unit Tests, for new Actual Bin Allocation Logic
# TODO: Organize & Fully comment the new Actual Bin Allocation Logic. Round in necessary places, and finalize final allocation messages
# TODO: Add TQDR Bars
# LK4Z3504J



#^###############################  Storage_Optimization.ipynb  ####################################
# ^ Author: Sukhendu Sain
# ^ Description: Main file of codebase. Houses main code for Storage Bin Allocation and Optimization
# ^ Data: 23-Nov-2024
#^################################################################################


# Import Necessary Libraries, Utils, and Config Files
from config import *
import pandas as pd
import numpy as np
import math
import openpyxl
from tqdm import tqdm
import utils

# # Data Import and Clean


####  AKINS Is OLD Data.  IGNORE AKINS Data for Processing
## @ Read FILE:: (AKINS FoMoCo_Piece_Sales_112222_YTD.xlsx) into Dataframe (Now Commented out)
# df_Akins = utils.read_excel(AKINS_FOMO_FILE_PATH)
# df_Akins['Part#'] = df_Akins['Part#'].apply(lambda a: "".join(str(a).split('-')))
# if print_df_after_import: utils.print_df(df_Akins, 200) # Print the Dataframe
# ~1-2secs



## @ Read FILE:: (GPARTS Part Measures.xlsx) into Dataframe
df_Gparts = utils.read_excel(GPARTS_FILE_PATH)
if print_df_after_import: utils.print_df(df_Gparts) # Print the Dataframe
# ~17 secs



## @ Read FILE:: (Wholesale JAN_Oct_Parts_Ranking_Counter_Invoices_All_Brands.xlsx) into Dataframe
df_Wholesale = utils.read_excel(WHOLESALE_FILE_PATH)

# Clean the Wholesale Dataframe
df_Wholesale['Description'] = df_Wholesale['Description'].astype(str)
df_Wholesale = df_Wholesale.drop(columns=[col for col in df_Wholesale.columns if 'Unnamed' in col], inplace=False)
df_Wholesale = df_Wholesale[(df_Wholesale['Vendor'] == 'FOR')].reset_index()
df_Wholesale.loc[df_Wholesale['Description'].apply(lambda x: len(x.split("      ")) > 1), 'Avg. Cost'] = df_Wholesale['Description'].apply(lambda x: [i for i in x.strip().split("      ")][-1])
df_Wholesale.loc[df_Wholesale['Description'].apply(lambda x: len(x.split("      ")) > 1), 'Description'] = df_Wholesale['Description'].apply(lambda x: "     ".join([i for i in x.strip().split("      ")][:-1]))

if print_df_after_import: utils.print_df(df_Wholesale) # Print the Dataframe
# ~7 secs



## @ Read FILE:: (Service JAN_Oct_Parts_Ranking_ROs_All_Brands.xlsx) into Dataframe
df_Service = utils.read_excel(SERVICE_FILE_PATH)

# Clean the Service Dataframe
df_Service['Description'] = df_Service['Description'].astype(str)
df_Service = df_Service.drop(columns=[col for col in df_Service.columns if 'Unnamed' in col], inplace=False)
df_Service = df_Service[(df_Service['Vendor'] == 'FOR')].reset_index()
df_Service.loc[df_Service['Description'].apply(lambda x: len(x.split("      ")) > 1), 'Avg. Cost'] = df_Service['Description'].apply(lambda x: [i for i in x.strip().split("      ")][-1])
df_Service.loc[df_Service['Description'].apply(lambda x: len(x.split("      ")) > 1), 'Description'] = df_Service['Description'].apply(lambda x: "     ".join([i for i in x.strip().split("      ")][:-1]))
df_Service.loc[df_Service['Qty Sold'].apply(lambda x: len(str(x).split("      ")) > 1), 'Dollars Sold'] = df_Service['Qty Sold'].apply(lambda x: [i for i in str(x).strip().split("      ")][-1])
df_Service.loc[df_Service['Qty Sold'].apply(lambda x: len(str(x).split("      ")) > 1), 'Qty Sold'] = df_Service['Qty Sold'].apply(lambda x: "     ".join([i for i in str(x).strip().split("      ")][:-1]))

if print_df_after_import: utils.print_df(df_Service, 100) # Print the Dataframe
# ~4 secs                                                                                                              



## @ Read FILE:: (Counter Pad) into Dataframe
df_CounterPad1 = utils.read_excel(COUNTERPAD_FILE_PATH, 0, None)
df_CounterPad2 = utils.read_excel(COUNTERPAD_FILE_PATH, 1, None)

df_CounterPad = pd.concat([df_CounterPad1, df_CounterPad2])

df_CounterPad.columns = df_CounterPad.iloc[1, :]
df_CounterPad = df_CounterPad.rename(columns={'Part #': 'Part#'}) # Rename the 'Part #' column to 'Part#'
df_CounterPad = df_CounterPad[(df_CounterPad['Vendor'] == 'FOR')].reset_index()

if print_df_after_import: utils.print_df(df_CounterPad, 100) # Print the Dataframe
# ~6 secs


# ## Data Processing & Calculation



# @ Make a Big Final Dataframe

# Make the Dataframe
df_Main = pd.DataFrame({
    'Part#': df_Gparts['Svc Part Number'],
    'Part Desc.': df_Gparts['Svc Part Number Description'],
    'Part Category': "",
    'Active': df_Gparts['Is Active?'],
    'Wholesale Sold': 0,
    'Service Sold': 0,
    'Total Sold': 0,
    'OH Inventory': 0,
    'SKU Count': 0,
    '0Dimensions': False,
    'Depth': df_Gparts['Prod Att - Length'],
    'Width': df_Gparts['Prod Att- Width'],
    'Height': df_Gparts['Prod Att - Height'],
    'Zone': "",
    'StorageType': "",
    'SubStorage': "",
    'Bin Type': "",
    'Num. Bin Required': 0,
    'Actual Bin Allocation': "",
    'Overflow Bins': "",
    'Overflow Comment': "",
    'Bin Location': ""
})

# df_Main = df_Gparts.loc[:, ['Svc Part Number', 'Svc Part Number Description', 'Is Active?', 'Prod Att - Length', 'Prod Att- Width', 'Prod Att - Height']]
# # for pn, pdesc, act, ln, wd, hg in zip(df_Gparts['Svc Part Number'], df_Gparts['Svc Part Number Description'], df_Gparts['Is Active?'], df_Gparts['Prod Att - Length'], df_Gparts['Prod Att- Width'], df_Gparts['Prod Att - Height']):
# #     df_Main.insert
# df_Main.columns = ['Part#', 'Part Desc.', 'Active', 'Depth', 'Width', 'Height']
# df_Main.insert(2, 'Part Category', None)
# df_Main.insert(4, 'Wholesale Sold', 0)
# df_Main.insert(5, 'Service Sold', 0)
# df_Main.insert(6, 'Total Sold', 0)
# df_Main.insert(7, 'OH Inventory', 0)
# df_Main.insert(8, 'SKU Count', 0)
# df_Main.insert(9, '0Dimensions', False)
# df_Main.insert(13, 'Zone', None)
# df_Main.insert(14, 'StorageType', None)
# df_Main.insert(15, 'SubStorage', None)
# df_Main.insert(16, 'Bin Type', None)
# df_Main.insert(17, 'Num. Bin Required', None)
# df_Main.insert(18, 'Actual Bin Allocation', None)
# df_Main.insert(19, 'Overflow Bins', None)
# df_Main.insert(20, 'Overflow Comment', None)
# df_Main.insert(21, 'Bin Location', None)


# Insert the Rows from other Files
oh_dict = df_Wholesale.set_index('Part Number')['Sold'].to_dict()
df_Main['Wholesale Sold'] = df_Main['Part#'].map(oh_dict)

oh_dict = df_Service.set_index('* indicates a superseded part\nPart Number')['Qty Sold'].to_dict()
df_Main['Service Sold'] = df_Main['Part#'].map(oh_dict)

oh_dict = df_CounterPad.set_index('Part#')['OH'].to_dict()
df_Main['OH Inventory'] = df_Main['Part#'].map(oh_dict)


# Process and Clean
# Drop rows with Sold and Inventory 'nan'
df_Main = df_Main.dropna(subset=["Wholesale Sold", "Service Sold", "OH Inventory"], how="all").reset_index(drop=True)
# Set 0Dimensions
df_Main.loc[(df_Main["Depth"] == 0) | (df_Main["Height"] == 0) | (df_Main["Width"] == 0), "0Dimensions"] = True
# Drop 0Dimensions Rows if drop0Dims
if drop0Dims: df_Main = df_Main[df_Main["0Dimensions"] == False]
# Set Total_Sold
df_Main["Total Sold"] = df_Main["Wholesale Sold"].astype(float) + df_Main["Service Sold"].astype(float)
# Sort by 'Total Sold'
df_Main = df_Main.sort_values('Total Sold', ascending=False)
# Remove Alphanumeric Strings
# Fill 'nan' with 0 and convert to float
df_Main['OH Inventory'] = pd.to_numeric(df_Main['OH Inventory'], errors='coerce')
df_Main['Wholesale Sold'] = df_Main['Wholesale Sold'].fillna(0).astype(float)
df_Main['Service Sold'] = df_Main['Service Sold'].fillna(0).astype(float)
df_Main['Total Sold'] = df_Main['Total Sold'].fillna(0).astype(float)
df_Main['OH Inventory'] = df_Main['OH Inventory'].fillna(0).astype(float)
# ^ Add Random Values for SKU Count temporarily
df_Main["SKU Count"] = np.random.choice(np.arange(20), size=len(df_Main), replace=True)

# ~0.4 secs



# from dateutil.parser import parse

# def is_non_numeric(x):
#     try:
#         # Try to parse as a number
#         float(x)
#         return False
#     except ValueError:
#         try:
#             # If it's not a number, try to parse as a date
#             parse(x)
#             return True
#         except ValueError:
#             # If it's neither a number nor a date, it's likely alphanumeric
#             return isinstance(x, str) and x.isalnum()

# # Apply the function to the 'OH Inventory' column
# df_Main['OH Inventory'] = df_Main['OH Inventory'].apply(lambda x: 0 if is_non_numeric(x) else x)

# # Convert the column to numeric, replacing non-numeric values with NaN
# df_Main['OH Inventory'] = pd.to_numeric(df_Main['OH Inventory'], errors='coerce')

# # Replace NaN values with 0
# df_Main['OH Inventory'] = df_Main['OH Inventory'].fillna(0)




utils.print_df(df_Main)
# ~12 secs



# * SAVE the Marged Main Dataset into Excel, To Avoid The Above Steps During Rerun
df_Main.to_excel('Gparts-MergedMainData.xlsx', index=False) 
# ~8 Secs


#####   READ  ALREADY SAVED GPARTS-Merged Data
## @ Read FILE:: (Gparts-MergedMainData.xlsx) into Dataframe
df_Main = utils.read_excel("Gparts-MergedMainData.xlsx").fillna("")
# ~6 secs



## ^ TESTING TIRE DATA  ---- Change some TIRE data Manually
df_Main.loc[df_Main["Part#"] == '6F2Z1A189A', ["Part Desc.","0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = ["6F2Z1A189A-TIRE",False, 28,28,3,100]
df_Main.loc[df_Main["Part#"] == '7L1Z1A189A', ["Part Desc.","0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = ["7L1Z1A189A-TIRE",False, 34,34,3.5,50]
df_Main.loc[df_Main["Part#"] == '9OO1183106436', ["0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = [False, 30,30,3,1000]
df_Main.loc[df_Main["Part#"] == '9OO439510', ["0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = [False, 45,45,4.5,500]
df_Main.loc[df_Main["Part#"] == '9OO1732002500', ["0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = [False, 50,50,5,250]
df_Main.loc[df_Main["Part#"] == '9OO3004901', ["0Dimensions", "Depth", "Height", "Width", "OH Inventory"]] = [False, 40,40,4,25]
# & df_Main[['TIRE' in s for s in df_Main["Part Desc."]]]



## ^ Part Categorization  BUT Client Will Share Actual Part Category
def part_categorization(df_toBeCategorized, categoryColName):
    # TODO: Add more categories
    for i in range(df_toBeCategorized.shape[0]):
        desc = "-".join(df_toBeCategorized.loc[i, "Part Desc."].split("-")[1:])
        pnum = df_toBeCategorized.loc[i, "Part#"][1:].upper()
        category = ""
        if "battery" in desc.lower():
            if desc.strip().lower() == "battery":
                category = "Battery"
            else:
                category = "Battery Accessory"
        elif "tire" in desc.lower():
            if desc.strip().lower() == "tire":
                category = "Tire"
            else:
                category = "Tire Accessory"            
        elif "hood" in desc.lower():
            if (desc.strip().lower() == "hood asy") | (desc.strip().lower() == "hood  asy"):
                category = "Hood"
            else:
                category = "Hood Accessory"
                
        elif (desc.strip().lower() == "cover"):
            category = "Bumper Cover"
        elif ("17D957" in pnum) | ("17K835" in pnum):
            category = "Bumper Cover"
        elif desc.strip().lower() == "seal":
            category = "Seal"
        elif desc.strip().lower() == "name plate":
            category = "Name Plate"
        elif desc.strip().lower() == "v-belt":
            category = "V-Belt"
        elif desc.strip().lower() == "v-belt":
            category = "V-Belt"
       
        elif ("blade" in desc.lower()) & ("wiper" in desc.lower()):
            category = "Wiper Blade"        
        elif ("arm" in desc.lower()) & ("wiper" in desc.lower()):
            category = "Wiper Arm"
        elif ("belt" in desc.lower()) & ("retractor" not in desc.lower()) & ("hole" not in desc.lower()) & ("cover" not in desc.lower()):
            category = "Belt"
        # elif ("hose" in desc.lower()) & ("vent" not in desc.lower()) & ("connect" not in desc.lower() & ("radiator" not in desc.lower()& ("heater" not in desc.lower()):
        #     category = "Hose"

        df_toBeCategorized.loc[i, categoryColName] = category
        



part_categorization(df_Main, 'Part Category')
# ~8 secs


# ### Apply Zoning



# @ Apply Zoning based on Time Period/Sale

## * Main Function for Apply Zoning
def Apply_Zoning(df_toBeZoned, zones, soldColName='Total Sold', zoneColName='Zone'): 
    df_toBeZoned.loc[:, zoneColName] = df_toBeZoned[soldColName].apply(lambda x: next((zone for zone, ratio in zones.items() if x >= ratio), list(zones.keys())[-1]))
    df_toBeZoned.loc[df_toBeZoned[soldColName] < 0, zoneColName] = None



## * Run the Apply_Zoning on df_Main
Apply_Zoning(df_Main, zones, 'Total Sold', 'Zone')


## * Check each Zone's number of Part Numbers
df_Main['Zone'].value_counts()





# ### Specialty Storage Assignment


# #### Function for Bin Calculation


def getNumOfBin(depth, width, height, raw_bin_dim, ohInven, fillFactor):
    # * Raw Bin Dimensions has this format :-  Height_Depth_Width
    if (raw_bin_dim != "") and (ohInven > 0):
        bin_height = float(raw_bin_dim.split("_")[1])
        bin_depth = float(raw_bin_dim.split("_")[2])
        bin_width = float(raw_bin_dim.split("_")[3])

        if raw_bin_dim.split("_")[0] == "BR":   # * Battery Rack
            return round((ohInven * width) / bin_depth, 3)      
               
        volBin = fillFactor * bin_height * bin_depth * bin_width    # * Available Storage Space
        volPart = height * depth * width
        if (volBin == 0):
            return 0
        
        numOfBins = round((ohInven * volPart) / volBin, 3)      # * Returns Fraction. 
        return numOfBins
    else:
        return 0


# ### Main Function for Storage Assignment


def getStorage(zone, pcate, depth, width, height, ohInven, fillFactor):
    # Initialize the empty Variables
    storageType = ""
    subStorage = ""

    if zone == "":
        return storageType, subStorage, "", 0 # Return the Values 

    isSpec, storageType, subStorage, raw_bin_dim = utils.getSpecialtyStorage(pcate, depth, width, height)

    if not isSpec: 
        if (zone == "Red Hot") | (zone == "Red"):
            storageType, subStorage, raw_bin_dim = utils.getRedHotStorage(depth, width, height)
        elif (zone == "Orange") | (zone == "Yellow"):
            storageType, subStorage, raw_bin_dim = utils.getOrangeYellowStorage(depth, width, height)
        elif (zone == "Green") | (zone == "Blue"):   
            storageType, subStorage, raw_bin_dim = utils.getGreenBlueStorage(depth, width, height)
 
    numOfBins = getNumOfBin(depth, width, height, raw_bin_dim, ohInven, fillFactor)
    binDim = ""

    # * Build Bin Label with C (Clip), B (Bulk), D (Drawer), Battery Rack (BR), Tire Rack (TR) and Width-Depth-Height
    if raw_bin_dim.strip():   
        binDim =  raw_bin_dim.split('_')[0] + raw_bin_dim.split('_')[3] + raw_bin_dim.split('_')[2] + raw_bin_dim.split('_')[1]
    

    return storageType, subStorage, binDim, numOfBins # Return the Values


# #### Apply the Storage Function


for i in tqdm(range(df_Main.shape[0]), desc="Completion"):
#for i in range(df_Main.shape[0]):
    # Set the Dimensions of the Data into Variables
    depth = df_Main.loc[i, "Depth"]
    height = df_Main.loc[i, "Height"]
    width = df_Main.loc[i, "Width"]

    zone = df_Main.loc[i, "Zone"]
    pcate = df_Main.loc[i, "Part Category"]
    ohInven = df_Main.loc[i, "OH Inventory"]

    # * If any dimension is zero, set empty Storage
    if df_Main.loc[i, "0Dimensions"] == True:
        df_Main.loc[i, "StorageType"] = ""
        df_Main.loc[i, "SubStorage"] = ""
        continue

    # Set Storage of the Parts
    df_Main.loc[i, "StorageType"], df_Main.loc[i, "SubStorage"], df_Main.loc[i, "Bin Type"], df_Main.loc[i, "Num. Bin Required"] = getStorage(zone, pcate, depth, width, height, ohInven, fillFactor)

# ~34 secs



## @ HANGING Storage Calculation
# TODO: To Get the SKU Count for Hanging Storage
# * ASSUMPTION: Hook Length Based on SKU Count:
# *             6-inch hooks: For SKUs with 10 items or fewer
# *             12-inch hooks: For SKUs with 10–20 items

for hangingPN in tqdm(df_Main.loc[df_Main['StorageType'] == 'Hanging Specialty Storage', "Part#"], "Completion"): # Get Hanging Parts
    # & df_Main.loc[(df_Main['Part#'] == hangingPN), "Num. Bin Required"] = round(int(df_Main.loc[(df_Main['Part#'] == hangingPN), "OH Inventory"].values[0]) / hookDiv, 4)
    df_Main.loc[(df_Main['Part#'] == hangingPN), "SubStorage"], df_Main.loc[(df_Main['Part#'] == hangingPN), "Bin Type"], hookDiv = ("6-inch Hook", "HS06", 10) if df_Main.loc[(df_Main['Part#'] == hangingPN), "SKU Count"].values[0] <= 10 else ("12-inch Hook", "HS12", 20)
    df_Main.loc[(df_Main['Part#'] == hangingPN), "Num. Bin Required"] = int(df_Main.loc[(df_Main['Part#'] == hangingPN), "OH Inventory"].values[0])  # * Set No. of Hooks = Inventory Count

###  SET SKU Count to zero for all other STORAGE Types 
df_Main.loc[df_Main['StorageType'] != 'Hanging Specialty Storage', "SKU Count"] = 0

# ~46 secs



## @ Tire Storage Calculation
# TODO: FINALIZE  the Calculation of Tire Carousel Model Selection Based on Percentage 
# * ASSUMPTION:  Assign Carousel Model based on Diameter Group %age
# *   If 50% of Tires Have 33-inch or More Diameter, Assign Large-Storage (72-width carousel)
# *   ELSE,  For 28-inch or less, and,  between 28-33 inches, assign standard carrousel (48-width carousel) 

carousel_model = 'TR72' if df_Main[df_Main['StorageType'] == 'Tire Specialty Storage'][df_Main['SubStorage'] == '33-inches Dia'].shape[0] / df_Main[df_Main['StorageType'] == 'Tire Specialty Storage'].shape[0] >= tirePercent else 'TR48'
carousel_width = 72 if df_Main[df_Main['StorageType'] == 'Tire Specialty Storage'][df_Main['SubStorage'] == '33-inches Dia'].shape[0] / df_Main[df_Main['StorageType'] == 'Tire Specialty Storage'].shape[0] >= tirePercent else 48
for tirePN in df_Main.loc[df_Main['StorageType'] == 'Tire Specialty Storage', "Part#"]:
    df_Main.loc[(df_Main['Part#'] == tirePN), "Num. Bin Required"] = round(int(df_Main.loc[(df_Main['Part#'] == tirePN), "OH Inventory"].values[0]) / (carousel_width // int(df_Main.loc[(df_Main['Part#'] == tirePN), "Width"].values[0])), 3)
    df_Main.loc[(df_Main['Part#'] == tirePN), "Bin Type"] = carousel_model

# ~0.3 secs



# @ Do Actual Storage Allocation based on Inventory, Number of Bins availiable, Handle Overflow, etc...

df_binData = pd.DataFrame(columns=['Bin Label', 'Bin Category', 'Total Bins', 'Filled Amount', 'Bin Order', 'GB Bin Order', 'Bin Location', 'Availiability Flag'])

# * High-Density Drawers (2)
binData = [ 
    {'Bin Label': 'D362406', 'Bin Category': 'Drawer', 'Total Bins': 5, 'Filled Amount': 0, 'Bin Order': 1, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
    {'Bin Label': 'D482406', 'Bin Category': 'Drawer', 'Total Bins': 4, 'Filled Amount': 0, 'Bin Order': 2, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'}
 ]
# * Clip-Shelving (6)
binData.extend([
    {'Bin Label': 'C361215', 'Bin Category': 'Clip', 'Total Bins': 4, 'Filled Amount': 0, 'Bin Order': 3, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
    {'Bin Label': 'C361815', 'Bin Category': 'Clip', 'Total Bins': 6, 'Filled Amount': 0, 'Bin Order': 4, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
    {'Bin Label': 'C362415', 'Bin Category': 'Clip', 'Total Bins': 8, 'Filled Amount': 0.0, 'Bin Order': 5, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},    
    {'Bin Label': 'C481215', 'Bin Category': 'Clip', 'Total Bins': 3, 'Filled Amount': 0.0, 'Bin Order': 6, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},  
    {'Bin Label': 'C481815', 'Bin Category': 'Clip', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 7, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},   
    {'Bin Label': 'C482415', 'Bin Category': 'Clip', 'Total Bins': 7, 'Filled Amount': 0.0, 'Bin Order': 8, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'}
])

# * Bulk-Storage (18)
binData.extend([
  {'Bin Label': 'B482448', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 9, 'GB Bin Order': 1,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B483648', 'Bin Category': 'Bulk', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 10, 'GB Bin Order': 2,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B484248', 'Bin Category': 'Bulk', 'Total Bins': 6, 'Filled Amount': 0.0, 'Bin Order': 11, 'GB Bin Order': 3,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B484848', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 12, 'GB Bin Order': 4,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B487248', 'Bin Category': 'Bulk', 'Total Bins': 3, 'Filled Amount': 0.0, 'Bin Order': 13, 'GB Bin Order': 5,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B489648', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 14, 'GB Bin Order': 6,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B722448', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 15, 'GB Bin Order': 7,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B723648', 'Bin Category': 'Bulk', 'Total Bins': 3, 'Filled Amount': 0.0, 'Bin Order': 16, 'GB Bin Order': 8,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B724248', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 17, 'GB Bin Order': 9,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B724848', 'Bin Category': 'Bulk', 'Total Bins': 6, 'Filled Amount': 0.0, 'Bin Order': 18, 'GB Bin Order': 10,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B727248', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 19, 'GB Bin Order': 11,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B729648', 'Bin Category': 'Bulk', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 20, 'GB Bin Order': 12,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B962448', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 21, 'GB Bin Order': 13,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B963648', 'Bin Category': 'Bulk', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 22, 'GB Bin Order': 14,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B964248', 'Bin Category': 'Bulk', 'Total Bins': 7, 'Filled Amount': 0.0, 'Bin Order': 23, 'GB Bin Order': 15,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B964848', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 24, 'GB Bin Order': 16,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B967248', 'Bin Category': 'Bulk', 'Total Bins': 6, 'Filled Amount': 0.0, 'Bin Order': 25, 'GB Bin Order': 17,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
  {'Bin Label': 'B969648', 'Bin Category': 'Bulk', 'Total Bins': 4, 'Filled Amount': 0.0, 'Bin Order': 26, 'GB Bin Order': 18,'Bin Location': 'None', 'Availiability Flag': 'Yes'}
])

# * Specialty (7) TR
binData.extend([
    {'Bin Label': 'BR484816', 'Bin Category': 'Battery', 'Total Bins': 14, 'Filled Amount': 0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
    {'Bin Label': carousel_model, 'Bin Category': 'Tire', 'Total Bins': 6, 'Filled Amount': 0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},
    {'Bin Label': 'BC967248', 'Bin Category': 'Bumper Cover', 'Total Bins': 3, 'Filled Amount': 0.0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},  
    {'Bin Label': 'BH967280', 'Bin Category': 'Hood', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},   
    {'Bin Label': 'HS06', 'Bin Category': 'Hanging', 'Total Bins': 5, 'Filled Amount': 0.0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'},   
    {'Bin Label': 'HS12', 'Bin Category': 'Hanging', 'Total Bins': 7, 'Filled Amount': 0.0, 'Bin Order': 0, 'GB Bin Order': 0,'Bin Location': 'None', 'Availiability Flag': 'Yes'}
])

# Append the Data to the DF
df_binData = pd.concat([df_binData, pd.DataFrame(binData)], ignore_index=True)
df_binData['Total Bins'] = np.random.choice(range(40, 400), df_binData.shape[0])


utils.print_df(df_binData)


# * Write  the Final BINS Dataset into Excel
#df_binData.to_excel('All_Bins_Data.xlsx', index=False) 



## ^ Psuedocode for main Logic
# ~ Start with 'Bin Type' as default assignment for the part --  ALLOCATE Bin As Per Availiability, Zone, and Priority Rules
# ~ Add a column to 'BINS' dataframe, as Order or priority based on dimension for a given Rack Type (Drawer, Bulk, Clip, etc). 
# ~  This Order should only within specific Rack Type (Drawer, Bulk, Clip, etc). To change to Rack Type, use priorities given for 
# ~ Zones and Part Category. Always start with minimum dimension of the BIN
# ~ Start with 'Red Hot' Zone Parts to allocate Actual Bin. 
# ~ LOOP Over Zones. Then LOOP Over Parts. For Each Part, Check default 'Bin Type'. If Available for storage, then allocate it  
# ~ (Add the Bin Type in 'Actual Bin Allocation',  Increment the Fill_Quantity for the 'Bin Type' in second dataframe for BINS. 
# ~ Fill_Quantity can be fraction.  If Fill_Quantity is same  as  Available Bins, then Set the 'Available FLAG' for the BIN to 'NO'
# ~ That means, the BIN is fully occupied. Pick the next available BIN. 
# ~ Think About Overflow,  and Bin Location (Not sure if Bin Locations will be zone-wise)



## @ Red Hot, Red, Orange, and Yellow Zone Non-Specialty Parts
for pn in tqdm(df_Main.loc[(df_Main['Zone'] == 'Red Hot') | (df_Main['Zone'] == 'Red') | (df_Main['Zone'] == 'Orange') | (df_Main['Zone'] == 'Yellow'), 'Part#'], "Completion"):
    # Get & Set Variables
    actualBin =  ""
    overflowBin = ""
    overflowComment = ""

    partData = df_Main[df_Main['Part#'] == pn]
    partHeight = partData['Height'].values[0]
    partWidth = partData['Width'].values[0]
    partDepth = partData['Depth'].values[0]
    partVolume = partHeight * partWidth * partDepth 
    binType = partData['Bin Type'].values[0]
    storageType = partData['StorageType'].values[0]
    partOHInven = partData['OH Inventory'].values[0]


    # ~ Base Continue Case - If no Storage Assignment, no Inventory Parts or is Specialty Part
    if (binType == "") or ("Specialty Storage" in storageType) or (partOHInven <= 0):     
        continue

    binData = df_binData.loc[df_binData['Bin Label'] == binType]
    totalBinOfType = binData['Total Bins'].values[0]
    filledAmtOfBin = binData['Filled Amount'].values[0]
    binOrder = binData['Bin Order'].values[0]
    flagAvail = binData['Availiability Flag'].values[0]

    # Calculate Variables
    binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
    remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
    partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

    if partsAllocated > 0:   # * If Actual Bin is Availiable
        # Calculate Variables
        totalPartVolume = partOHInven * partVolume
        actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
        numBins = round(totalPartVolume / binVolume, 3)

        # Set Values
        df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
        actualBin = f"{binType} ({actualBinRequired}, {partsAllocated})"

        if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): # If Bin is FULL, Reset Flag
            df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

        if (filledAmtOfBin + numBins) > (totalBinOfType- 0.01):   # * If Overflow (Actual Bin can't fit all Parts) (Only 1)
            overflowParts = partOHInven - partsAllocated
            
            # Find next availiable BIN to  handle overflow parts
            binFound = False
            for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
                # Calculate Variables
                binData = df_binData[df_binData['Bin Label'] == binType1]
                totalBinOfType = binData['Total Bins'].values[0]
                filledAmtOfBin = binData['Filled Amount'].values[0]
                binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
                remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
                partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

                if partsAllocated > 0:   # If Actual Bin is Availiable
                    binFound = True
                    break
                #  OLD Code
                # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
                #     binData = df_binData[df_binData['Bin Label'] == binType1]
                #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
                #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
                #     if (fillAmt < (totalBin - 0.01)):
                #         break

            if not (binFound):
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part1- No More Available/Fitting Bins"
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
                continue

            # Get & Calculate Variables
            binType = binData['Bin Label'].values[0]
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            
            binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
            actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)

            # Add  FilledAmount for OverFlow Bin  
            df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
            overflowBin = f"{binType} ({actualBinRequired}, {partsAllocated})"
        
            # If more overflow
            leftParts = overflowParts - partsAllocated
            if leftParts > 0:
                binsNeeded = round(leftParts * partVolume / binVolume, 3)
                overflowComment = f"Second Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"
            
            if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): 
                df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

    else:    # * If Actual Bin is not avaialble
        binFound = False
        for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
            # Calculate Variables
            binData = df_binData[df_binData['Bin Label'] == binType1]
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

            if partsAllocated > 0:
                binFound = True
                break
            #  OLD Code
            # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
            #     binData = df_binData[df_binData['Bin Label'] == binType1]
            #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
            #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
            #     if (fillAmt < (totalBin - 0.01)):
            #         break

        if not (binFound):
            df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = "Part2- No More Available/Fitting Bins"
            df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = "Part2- No More Available/Fitting Bins"
            df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{partOHInven} Parts Left; No Bins availiable to fit"
            continue

        binType = binData['Bin Label'].values[0]
        totalBinOfType = binData['Total Bins'].values[0]
        filledAmtOfBin = binData['Filled Amount'].values[0]
        binOrder = binData['Bin Order'].values[0]
        binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
        remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume   ## Check for Remaining Vol in Bin 
        partVolume = partHeight * partWidth * partDepth 
        totalPartVolume = partOHInven * partVolume
        numBins = round(totalPartVolume / binVolume, 3)         ### Number of Bins Required to fill Inventry Parts
        partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)
        actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
        
        df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
        actualBin = f"{binType} ({actualBinRequired}, {partsAllocated})"

        if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): 
            df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

        if filledAmtOfBin + numBins > totalBinOfType:   # * If Overflow (Actual Bin can't fit all Parts) (Only 1)
            overflowParts = partOHInven - partsAllocated
            for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
                # Calculate Variables
                binData = df_binData[df_binData['Bin Label'] == binType1]
                totalBinOfType = binData['Total Bins'].values[0]
                filledAmtOfBin = binData['Filled Amount'].values[0]
                binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
                remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
                partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

                if partsAllocated > 0:
                    binFound = True
                    break
                #  OLD Code
                # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
                #     binData = df_binData[df_binData['Bin Label'] == binType1]
                #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
                #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
                #     if (fillAmt < (totalBin - 0.01)):
                #         break

            if not (binFound):
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part3- No More Available/Fitting Bins"
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
                continue

            binType = binData['Bin Label'].values[0]
            binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
            actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
            
            df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
            overflowBin = f"{binType} ({actualBinRequired}, {partsAllocated})"
        
            leftParts = overflowParts - partsAllocated
            if leftParts > 0:
                binsNeeded = round(leftParts * partVolume / binVolume, 2)
                overflowComment = f"Second Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"

            if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): 
                df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

    df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = actualBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = overflowBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = overflowComment

# ~1 min 15 secs



## @ Green and Blue Zone Non-Specialty Parts
for pn in tqdm(df_Main.loc[(df_Main['Zone'] == 'Green') | (df_Main['Zone'] == 'Blue'), 'Part#'], "Completion"):
    # Get Variable
    actualBin =  ""
    overflowBin = ""
    overflowComment = ""

    partData = df_Main[df_Main['Part#'] == pn]
    partHeight = partData['Height'].values[0]
    partWidth = partData['Width'].values[0]
    partDepth = partData['Depth'].values[0]
    binType = partData['Bin Type'].values[0]
    storageType = partData['StorageType'].values[0]
    partOHInven = partData['OH Inventory'].values[0]

    ## * If Inventry Qty is zero, Go to Next Part -- Handle Specialty Storage separately 
    if (binType == "") or ("Specialty Storage" in storageType) or (partOHInven <= 0):     
        continue

    binData = df_binData.loc[df_binData['Bin Label'] == binType]
    totalBinOfType = binData['Total Bins'].values[0]
    filledAmtOfBin = binData['Filled Amount'].values[0]
    binOrder = binData['GB Bin Order'].values[0]
 
    # Calculate Variables
    binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
    remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
    partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

    if partsAllocated > 0: # & (binData['Availiability Flag'].values[0] == "Yes"):    # If Actual Bin is Availiable
        binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
        remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
        partVolume = partHeight * partWidth * partDepth 
        totalPartVolume = partOHInven * partVolume
        numBins = round(totalPartVolume / binVolume, 3)
        partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)
        binRequired = round((partsAllocated * partVolume) / binVolume, 3)    

        df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += binRequired
        actualBin = f"{binType} ({binRequired}, {partsAllocated})"

        if filledAmtOfBin + ((partsAllocated * partVolume) / binVolume) >= (totalBinOfType - 0.01):    ### @ ACtual BIN
            df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

        ###  Handle OVERFLOW  condition (Only 1 Overflow) 
        if filledAmtOfBin + numBins > totalBinOfType:
            
            overflowParts = partOHInven - partsAllocated
            for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
                # Calculate Variables
                binData = df_binData[df_binData['Bin Label'] == binType1]
                totalBinOfType = binData['Total Bins'].values[0]
                filledAmtOfBin = binData['Filled Amount'].values[0]
                binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
                remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
                partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

                if partsAllocated > 0:   # @ If Actual Bin is Availiable
                    binFound = True
                    break
                #  OLD Code
                # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
                #     binData = df_binData[df_binData['Bin Label'] == binType1]
                #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
                #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
                #     if (fillAmt < (totalBin - 0.01)):
                #         break

            if not (binFound):
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part1-No More Available/Fitting Bins"
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
                continue

            binType = binData['Bin Label'].values[0]
            binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
            binRequired = round((partsAllocated * partVolume) / binVolume, 3)  

            overflowBin = f"{binType} ({binRequired}, {partsAllocated})"
        
            leftParts = overflowParts - partsAllocated
            if leftParts > 0:
                binsNeeded = round(leftParts * partVolume / binVolume, 3)
                overflowComment = f"Second Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"
            
            if (filledAmtOfBin + binRequired) >= (totalBinOfType - 0.01): 
                df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

    else:    ## ^  If  suggested Bin Is NOT Avaialble.  Pick Next Available Bin and Process
        for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
            # Calculate Variables
            binData = df_binData[df_binData['Bin Label'] == binType1]
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

            if partsAllocated > 0:   # @ If Actual Bin is Availiable
                binFound = True
                break
            #  OLD Code
            # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
            #     binData = df_binData[df_binData['Bin Label'] == binType1]
            #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
            #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
            #     if (fillAmt < (totalBin - 0.01)):
            #         break

        if not (binFound):            
            df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = "Part2-No More Available/Fitting Bins"
            df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = "Part2-No More Available/Fitting Bins"
            df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{partOHInven} Parts Left; No Bins availiable to fit"
            continue

        binType = binData['Bin Label'].values[0]
        totalBinOfType = binData['Total Bins'].values[0]
        filledAmtOfBin = binData['Filled Amount'].values[0]
        binOrder = binData['GB Bin Order'].values[0]
        binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
        remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume   ## Check for Remaining Vol in Bin 
        partVolume = partHeight * partWidth * partDepth 
        totalPartVolume = partOHInven * partVolume
        numBins = round(totalPartVolume / binVolume, 3)         ### Number of Bins Required to fill Inventry Parts
        partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)
        binRequired = round((partsAllocated * partVolume) / binVolume, 3)  
        
        df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += binRequired
        actualBin = f"{binType} ({binRequired}, {partsAllocated})"

        if (filledAmtOfBin + binRequired) >= (totalBinOfType - 0.01):
            df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

        ###  Handle OVERFLOW  condition (Only 1 Overflow) 
        if filledAmtOfBin + numBins > totalBinOfType:
            overflowParts = partOHInven - partsAllocated
            for binType1 in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
                # Calculate Variables
                binData = df_binData[df_binData['Bin Label'] == binType1]
                totalBinOfType = binData['Total Bins'].values[0]
                filledAmtOfBin = binData['Filled Amount'].values[0]
                binVolume = fillFactor * (float(binType1[1:3]) * float(binType1[3:5]) * float(binType1[5:7]))
                remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
                partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

                if partsAllocated > 0:   # @ If Actual Bin is Availiable
                    binFound = True
                    break
                #  OLD Code
                # if (df_binData[df_binData['Bin Label'] == binType1]['Availiability Flag'].values[0] == 'Yes'):
                #     binData = df_binData[df_binData['Bin Label'] == binType1]
                #     fillAmt = df_binData[df_binData['Bin Label'] == binType1]['Filled Amount'].values[0] 
                #     totalBin = df_binData[df_binData['Bin Label'] == binType1]['Total Bins'].values[0] 
                #     if (fillAmt < (totalBin - 0.01)):
                #         break

            if not (binFound):
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part3-No More Available/Fitting Bins"
                df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
                continue

            binType = binData['Bin Label'].values[0]
            binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
            totalBinOfType = binData['Total Bins'].values[0]
            filledAmtOfBin = binData['Filled Amount'].values[0]
            remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
            partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
            binRequired = round((partsAllocated * partVolume) / binVolume, 3) 

            ## @ ADD Below Filled Amount for Overflow bins
            df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += binRequired
            overflowBin = f"{binType} ({binRequired}, {partsAllocated})"
        
            leftParts = overflowParts - partsAllocated
            if leftParts > 0:
                binsNeeded = round(leftParts * partVolume / binVolume, 3)
                overflowComment = f"Second Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"

            if (filledAmtOfBin + binRequired) >= (totalBinOfType - 0.01):
                df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

    df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = actualBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = overflowBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = overflowComment

    
# ~16 mins 12 secs



## @ All Zone Specialty Parts
for pn in df_Main['Part#']:
    # Get & Set Variables
    actualBin =  ""
    overflowBin = ""
    overflowComment = ""

    partData = df_Main[df_Main['Part#'] == pn]
    partHeight = partData['Height'].values[0]
    partWidth = partData['Width'].values[0]
    partDepth = partData['Depth'].values[0]
    binType = partData['Bin Type'].values[0]
    partOHInven = partData['OH Inventory'].values[0]
    partSKUCount = partData['SKU Count'].values[0]

    # ~ Base Continue Case - If no Storage Assignment, no Inventory Parts or is not Specialty Part
    if (binType == "") or (partOHInven == 0):     
        continue
    if (all([binLabelTypes not in binType.lower() for binLabelTypes in ['br', 'tr', 'hs', 'bc', 'bh']])):     
        continue

    binData = df_binData.loc[df_binData['Bin Label'] == binType]
    totalBinOfType = binData['Total Bins'].values[0]
    filledAmtOfBin = binData['Filled Amount'].values[0]

    ## * Calculation for Battery
    if "br" in binType.lower():
        remainingBinWidth = (totalBinOfType - filledAmtOfBin) * float(binType[4:6])
        partsAllocated = min(math.floor(remainingBinWidth / partWidth), partOHInven)
        binsRequired = round((partsAllocated * partWidth) / float(binType[4:6]), 4)
        partsLeft = partOHInven - partsAllocated
        overflowBinsNeeded = partsLeft / float(binType[2:4])
    ## * Calculation for Tire
    if "tr" in binType.lower():
        remainingBinWidth = (totalBinOfType - filledAmtOfBin) * float(binType[2:4])
        partsAllocated = min(math.floor(remainingBinWidth / partWidth), partOHInven)
        binsRequired = round((partsAllocated * partWidth) / float(binType[2:4]), 4)
        partsLeft = partOHInven - partsAllocated
        overflowBinsNeeded = round(partsLeft / float(binType[2:4]), 3)
    ## * Calculation for Hanging Storage
    if "hs" in binType.lower():
        partsAllocated = min(math.floor(totalBinOfType - filledAmtOfBin), partOHInven)
        binsRequired = round(partsAllocated, 4)
        partsLeft = partOHInven - partsAllocated
        overflowBinsNeeded = partsLeft
    ## * Calculation for Bumper Cover & Hoods
    if ("bc" in binType.lower()) | ("bh" in binType.lower()):
        partVolume = partHeight * partWidth * partDepth 
        binVol = fillFactor * (float(binType[2:4]) * float(binType[4:6]) * float(binType[6:8]))
        partsAllocated = min(math.floor(((totalBinOfType - filledAmtOfBin) * binVol) / partVolume), partOHInven)
        binsRequired = round((partsAllocated * partVolume) / binVol, 4)
        partsLeft = partOHInven - partsAllocated
        overflowBinsNeeded = round((partsLeft * partVolume) / binVolume, 3)

    ## * Main Data Update
    if filledAmtOfBin + binsRequired >= (totalBinOfType - 0.01): # Update Availiability Flag, if full
        df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

    if partsAllocated > 0: # If any parts allocated
        df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += binsRequired
        actualBin = f"{binType} ({binsRequired}, {partsAllocated})"
    else:
        actualBin = "SP- No More Available/Fitting Bins" # If no parts can be allocated

    if partsLeft > 0: # If Overflow
        overflowBin = "SP- No More Available/Fitting Bins"
        overflowComment = f"{partOHInven - partsAllocated} Parts Left; {overflowBinsNeeded} quantity of {binType} Bin Needed;"

    df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = actualBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = overflowBin
    df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = overflowComment
           

# ~ 11 mins 28 sec


df_Main



# * Write  the Final Dataset into Excel
df_Main.to_excel('Final_Dataset.xlsx', index=False) 

# * Write  the Final BINS Dataset into Excel
df_binData.to_excel('All_Bins_Data.xlsx', index=False)


wb = openpyxl.load_workbook('Final_Dataset.xlsx')
ws = wb.active

# Adjust column widths
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 52
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 0
ws.column_dimensions['E'].width = 7
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 13
ws.column_dimensions['J'].width = 11
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 9
ws.column_dimensions['M'].width = 9
ws.column_dimensions['N'].width = 9
ws.column_dimensions['O'].width = 10
ws.column_dimensions['P'].width = 24
ws.column_dimensions['Q'].width = 34
ws.column_dimensions['R'].width = 12
ws.column_dimensions['S'].width = 18
ws.column_dimensions['T'].width = 33
ws.column_dimensions['U'].width = 32
ws.column_dimensions['V'].width = 46
ws.column_dimensions['W'].width = 12

ws.column_dimensions['D'].hidden = True

ws.freeze_panes = 'A2'

for cell in ws['B']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['E']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['O']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['P']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['Q']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['R']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['S']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['T']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['U']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for cell in ws['V']:
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


wb.save('FinalDataset.xlsx')


# ## * Main Loop for All Parts 
# for pn in df_Main['Part#']:
#     # Get & Set Variables
#     actualBin =  ""
#     overflowBin = ""
#     overflowComment = ""

#     partData = df_Main[df_Main['Part#'] == pn]
#     partHeight = partData['Height'].values[0]
#     partWidth = partData['Width'].values[0]
#     partDepth = partData['Depth'].values[0]
#     partVolume = partHeight * partWidth * partDepth 
#     binType = partData['Bin Type'].values[0]
#     partOHInven = partData['OH Inventory'].values[0]
#     totalPartVolume = partOHInven * partVolume

#     # ~ Base Continue Case - If no Storage Assignment, no Inventory Parts or is Specialty Part
#     if (binType == "") or (any([binLabelTypes in binType.lower() for binLabelTypes in ['br', 'tr', 'hs', 'bc', 'bh']])) or (partOHInven == 0):     
#         continue

#     binData = df_binData.loc[df_binData['Bin Label'] == binType]
#     totalBinOfType = binData['Total Bins'].values[0]
#     filledAmtOfBin = binData['Filled Amount'].values[0]
#     binOrder = binData['Bin Order'].values[0]

#     remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
#     partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)

#     if partsAllocated > 0:   # * If Actual Bin is Availiable
#         # Calculate Variables
#         binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
#         actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
#         numBins = round(totalPartVolume / binVolume, 3)
        
#         # Set Values
#         df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
#         actualBin = f"{binType} ({actualBinRequired}, {partsAllocated})"

#         if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): # Update Availiability Flag, if full
#             df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

#         if filledAmtOfBin + numBins > totalBinOfType:   # * If Overflow (Actual Bin can't fit all Parts) (Only 1)
#             overflowParts = partOHInven - partsAllocated
            
#             # Find next suitable availiable bin to fill overflow parts
#             for newBinType in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:   
#                 if (df_binData[df_binData['Bin Label'] == newBinType]['Availiability Flag'].values[0] == 'Yes'):
#                     binData = df_binData[df_binData['Bin Label'] == newBinType]
#                     break

#             if (binType == binData['Bin Label'].values[0]): # If no bins found
#                 df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part1 - No More Available/Fitting Bins"
#                 df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
#                 continue

#             # Get & Calculate Variables
#             binType = binData['Bin Label'].values[0]
#             totalBinOfType = binData['Total Bins'].values[0]
#             filledAmtOfBin = binData['Filled Amount'].values[0]
            
#             binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
#             remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
#             partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
#             actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)

#             # Set Values -- Add Filled Amount & Build overflowBin string
#             df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
#             overflowBin = f"{binType} ({actualBinRequired}, {partsAllocated})"
        
#             # If more overflow
#             leftParts = overflowParts - partsAllocated
#             if leftParts > 0:
#                 binsNeeded = round(leftParts * partVolume / binVolume, 3)
#                 overflowComment = f"More Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"
            
#             if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): # Update Availiability Flag, if full
#                 df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

#     else:   # * If Main Bin not Availiable - Pick next suitable availiable Bin
#         for newBinType in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
#             if (df_binData[df_binData['Bin Label'] == newBinType]['Availiability Flag'].values[0] == 'Yes'):
#                 binData = df_binData[df_binData['Bin Label'] == newBinType]
#                 #  & (df_binData[df_binData['Bin Label'] == binType]["Bin Label"][1:3].isdigit())
#                 break
                
#         if (binType == binData['Bin Label'].values[0]):
#             df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = "Part2 - No More Available/Fitting Bins"
#             df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = "No More Available/Fitting Bins"
#             df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{partOHInven} Parts Left; No Bins availiable to fit"
#             continue

#         binType = binData['Bin Label'].values[0]
#         totalBinOfType = binData['Total Bins'].values[0]
#         filledAmtOfBin = binData['Filled Amount'].values[0]
#         binOrder = binData['Bin Order'].values[0]
#         binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
#         remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume   ## Check for Remaining Vol in Bin 
#         partVolume = partHeight * partWidth * partDepth 
#         totalPartVolume = partOHInven * partVolume
#         numBins = round(totalPartVolume / binVolume, 2)
#         partsAllocated = min(math.floor(remainingBinVolume / partVolume), partOHInven)
#         actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
        
#         df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
#         actualBin = f"{binType} ({actualBinRequired}, {partsAllocated})"

#         if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): # Update Availiability Flag, if full
#             df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

#         # * If Overflow (Actual Bin can't fit all Parts) (Only 1)
#         if filledAmtOfBin + numBins > totalBinOfType:
#             overflowParts = partOHInven - partsAllocated
#             for binType in df_binData.loc[(df_binData['Bin Order'] > binOrder), 'Bin Label']:
#                 if (df_binData[df_binData['Bin Label'] == binType]['Availiability Flag'].values[0] == 'Yes'):
#                     binData = df_binData[df_binData['Bin Label'] == binType]
#                     break

#             if (binType == binData['Bin Label'].values[0]):
#                 df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = f"Part3- No More Available/Fitting Bins"
#                 df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = f"{overflowParts} Parts Left; No Bins availiable to fit"
#                 continue

#             binType = binData['Bin Label'].values[0]
#             binVolume = fillFactor * (float(binType[1:3]) * float(binType[3:5]) * float(binType[5:7]))
#             totalBinOfType = binData['Total Bins'].values[0]
#             filledAmtOfBin = binData['Filled Amount'].values[0]
#             remainingBinVolume = float(totalBinOfType - filledAmtOfBin) * binVolume  # Check for Remaining Vol in Bin 
#             partsAllocated = min(math.floor(remainingBinVolume / partVolume), overflowParts)
#             actualBinRequired = round((partsAllocated * partVolume) / binVolume, 3)
            
#             # Set Values -- Add Filled Amount & Build overflowBin string
#             df_binData.loc[df_binData['Bin Label'] == binType, 'Filled Amount'] += actualBinRequired
#             overflowBin = f"{binType} ({actualBinRequired}, {partsAllocated})"
        
#             leftParts = overflowParts - partsAllocated
#             if leftParts > 0:
#                 binsNeeded = round(leftParts * partVolume / binVolume, 2)
#                 overflowComment = f"Second Overflow: {leftParts} Parts Left; {binsNeeded} quantity of {binType} Bin Needed;"

#             if filledAmtOfBin + actualBinRequired >= (totalBinOfType - 0.01): 
#                 df_binData.loc[df_binData['Bin Label'] == binType, 'Availiability Flag'] = 'No'

#     df_Main.loc[df_Main['Part#'] == pn, 'Actual Bin Type'] = actualBin
#     df_Main.loc[df_Main['Part#'] == pn, 'Overflow Bins'] = overflowBin
#     df_Main.loc[df_Main['Part#'] == pn, 'Overflow Comment'] = overflowComment

# ~ 20 sec


